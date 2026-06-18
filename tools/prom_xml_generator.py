#!/usr/bin/env python3
"""
tools/prom_xml_generator.py — універсальний генератор Prom XML для авторозбірки.

Запуск:
  python3 tools/prom_xml_generator.py \
    --file /home/tekken/Downloads/catalog.xlsx \
    --category "Шланг системы охлаждения" \
    --portal-cat-id 13100650 \
    --name-ua "Шланг системи охолодження" \
    --out exports/prom_shlang_ohlagennya.xml
"""

import argparse
import os
import sys
from html import escape
from pathlib import Path

from openpyxl import load_workbook

# ── індекси колонок (0-based) ─────────────────────────────────────────────────
COL_ID         = 0
COL_PHOTO      = 1
COL_PART       = 2
COL_BRAND      = 3   # Марка авто
COL_MODEL      = 4   # Модель авто
COL_BODY       = 5   # Кузов
COL_ENGINE     = 6   # Двигатель
COL_YEAR       = 7   # Год выпуска
COL_FRONT_REAR = 8   # Передний/Задний
COL_LEFT_RIGHT = 9   # Левый/Правый
COL_PRICE      = 11  # Цена
COL_COLOR      = 13  # Цвет
COL_COMMENT    = 14  # Комментарий
COL_VENDOR     = 15  # Производитель (виробник деталі)
COL_PART_NUM   = 16  # Номер производителя
COL_CROSS      = 17  # Кросс-номера
COL_AVAIL      = 21  # Украина (свободно)

# ── виправлення брендів ───────────────────────────────────────────────────────
BRAND_FIX = {
    "aud":       "Audi",
    "volkswgen": "Volkswagen",
    "vag":       "Volkswagen",
    "china":     "Без бренда",
}


def fix_brand(val: str) -> str:
    if not val:
        return val
    fixed = BRAND_FIX.get(val.strip().lower())
    return fixed if fixed else val.strip()


def cell(row: tuple, idx: int, default: str = "") -> str:
    if idx >= len(row) or row[idx] is None:
        return default
    return str(row[idx]).strip()


def build_name(name_ua: str, brand: str, model: str, year: str,
               left_right: str, part_num: str = "", max_len: int = 110) -> str:
    parts = [p for p in [name_ua, brand, model, year, left_right] if p]
    name = " ".join(parts)
    if part_num:
        name = f"{name} {part_num}"
    return name[:max_len]


def build_description(row: tuple) -> str:
    fields = [
        ("Марка",            fix_brand(cell(row, COL_BRAND))),
        ("Модель",           cell(row, COL_MODEL)),
        ("Рік випуску",      cell(row, COL_YEAR)),
        ("Двигун",           cell(row, COL_ENGINE)),
        ("Кузов",            cell(row, COL_BODY)),
        ("Передній/Задній",  cell(row, COL_FRONT_REAR)),
        ("Лівий/Правий",     cell(row, COL_LEFT_RIGHT)),
        ("Колір",            cell(row, COL_COLOR)),
        ("Виробник",         cell(row, COL_VENDOR)),
        ("Номер виробника",  cell(row, COL_PART_NUM)),
        ("Кросс-номери",     cell(row, COL_CROSS)),
        ("Коментар",         cell(row, COL_COMMENT)),
    ]
    parts = [f"{label}: {val}" for label, val in fields if val]
    return ". ".join(parts) + ". Б/В запчастина зі розбірки."


def x(val: str) -> str:
    return escape(val, quote=False)


def generate_xml(rows: list, category_name: str, cat_id_local: int,
                 portal_cat_id: int, name_ua: str, no_filter: bool = False) -> tuple[str, dict]:

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<yml_catalog>',
        '  <shop>',
        '    <categories>',
        f'      <category id="{cat_id_local}">{x(category_name)}</category>',
        '    </categories>',
        '    <offers>',
    ]

    is_door = "двер" in name_ua.lower()

    stats = {
        "total_in":    0,
        "skip_photo":  0,
        "skip_avail":  0,
        "skip_filter": 0,
        "generated":   0,
    }

    for row in rows:
        stats["total_in"] += 1

        # фото
        photo_raw = cell(row, COL_PHOTO)
        photos = [p.strip() for p in photo_raw.split(",") if p.strip()]
        if not photos:
            stats["skip_photo"] += 1
            continue

        # наявність
        try:
            avail = int(float(row[COL_AVAIL] or 0)) if COL_AVAIL < len(row) else 0
        except (ValueError, TypeError):
            avail = 0
        if avail == 0:
            stats["skip_avail"] += 1
            continue

        # поля
        offer_id   = cell(row, COL_ID)
        brand      = fix_brand(cell(row, COL_BRAND))
        model      = cell(row, COL_MODEL)
        year       = cell(row, COL_YEAR)
        engine     = cell(row, COL_ENGINE)
        body       = cell(row, COL_BODY)
        front_rear = cell(row, COL_FRONT_REAR)
        left_right = cell(row, COL_LEFT_RIGHT)
        vendor     = cell(row, COL_VENDOR)
        part_num   = cell(row, COL_PART_NUM)
        cross      = cell(row, COL_CROSS)

        try:
            price = int(float(row[COL_PRICE] or 0)) if COL_PRICE < len(row) else 0
        except (ValueError, TypeError):
            price = 0

        oem = cell(row, COL_PART_NUM)
        if not no_filter and (not oem or price <= 1):
            stats["skip_filter"] += 1
            continue

        name = build_name(name_ua, brand, model, year, left_right, part_num)
        desc = build_description(row)

        lines.append(f'      <offer id="{x(offer_id)}" available="true" type="vendor.model">')
        lines.append(f'        <article>{x(str(offer_id))}</article>')
        lines.append(f'        <name>{x(name)}</name>')
        lines.append(f'        <name_ua>{x(name)}</name_ua>')
        photos = photos[:10]
        for photo in photos:
            lines.append(f'        <picture>{x(photo)}</picture>')
        lines.append(f'        <price>{price}</price>')
        lines.append(f'        <currencyId>USD</currencyId>')
        lines.append(f'        <categoryId>{cat_id_local}</categoryId>')
        lines.append(f'        <portal_category_id>{portal_cat_id}</portal_category_id>')
        if vendor:
            lines.append(f'        <vendor>{x(vendor)}</vendor>')
        model_val = part_num if part_num else model
        if model_val:
            lines.append(f'        <model>{x(model_val)}</model>')
        if part_num:
            lines.append(f'        <vendorCode>{x(part_num)}</vendorCode>')
        lines.append(f'        <description>{x(desc)}</description>')
        lines.append(f'        <description_ua>{x(desc)}</description_ua>')
        lines.append(f'        <selling_type>retail</selling_type>')
        lines.append(f'        <condition>used</condition>')

        # params
        if engine:
            lines.append(f'        <param name="Двигун">{x(engine)}</param>')
        if body:
            lines.append(f'        <param name="Кузов">{x(body)}</param>')
        if front_rear:
            lines.append(f'        <param name="Передній/Задній">{x(front_rear)}</param>')
        if left_right:
            lines.append(f'        <param name="Лівий/Правий">{x(left_right)}</param>')
        if cross:
            lines.append(f'        <param name="Кросс-номери">{x(cross)}</param>')
        if part_num:
            lines.append(f'        <param name="Код запчастини">{x(part_num)}</param>')
        if brand:
            lines.append(f'        <param name="Сумісність з маркою">{x(brand)}</param>')
        if model:
            lines.append(f'        <param name="Сумісність з моделлю">{x(model)}</param>')
        if year:
            lines.append(f'        <param name="Рік випуску автомобіля">{x(year)}</param>')
        lines.append(f'        <param name="Стан">Б/в</param>')

        if is_door:
            front_back = "Передня" if "Передн" in front_rear else "Задня"
            if "Лев" in left_right:
                side = "ліва"
            elif "Прав" in left_right:
                side = "права"
            else:
                side = ""
            location = f"{front_back} {side}".strip() if side else front_back
            lines.append(f'        <param name="Місце встановлення">{x(location)}</param>')
            lines.append(f'        <param name="Фарбування">Потрібно</param>')
            color = cell(row, COL_COLOR)
            if color:
                lines.append(f'        <param name="Колір">{x(color)}</param>')

        lines.append(f'      </offer>')

        stats["generated"] += 1

    lines += ['    </offers>', '  </shop>', '</yml_catalog>']
    return "\n".join(lines), stats


def main():
    parser = argparse.ArgumentParser(description="Prom XML generator для авторозбірки")
    parser.add_argument("--file",          required=True, help="Вхідний XLSX файл")
    parser.add_argument("--category",      required=True, help="Запчасть (рос.) для фільтру col[2]")
    parser.add_argument("--portal-cat-id", required=True, type=int, dest="portal_cat_id",
                        help="ID категорії Prom.ua (portal_category_id)")
    parser.add_argument("--name-ua",       required=True, dest="name_ua",
                        help="Назва категорії укр. (для заголовку товару)")
    parser.add_argument("--out",           required=True, help="Вихідний XML файл")
    parser.add_argument("--no-filter",     action="store_true", dest="no_filter", help="Без фільтру OEM/ціни")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        sys.exit(f"Файл не знайдено: {args.file}")

    print(f"Читаємо: {args.file}")
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    print(f"Всього рядків у файлі: {len(all_rows) - 1}")

    cat_lower = args.category.strip().lower()
    filtered = [
        r for r in all_rows[1:]
        if r[COL_PART] is not None and str(r[COL_PART]).strip().lower() == cat_lower
    ]
    print(f"Знайдено '{args.category}': {len(filtered)} рядків")

    if not filtered:
        # показати унікальні категорії для діагностики
        cats = sorted({str(r[COL_PART]).strip() for r in all_rows[1:]
                       if r[COL_PART] and args.category[:5].lower() in str(r[COL_PART]).lower()})
        if cats:
            print("Схожі категорії в файлі:")
            for c in cats[:10]:
                print(f"  {c}")
        sys.exit("Нічого не знайдено — перевірте --category")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    xml_content, stats = generate_xml(
        rows=filtered,
        category_name=args.name_ua,
        cat_id_local=1,
        portal_cat_id=args.portal_cat_id,
        name_ua=args.name_ua,
        no_filter=args.no_filter,
    )

    out_path.write_text(xml_content, encoding="utf-8")

    print()
    print("─" * 50)
    print(f"Знайдено в категорії          : {stats['total_in']}")
    print(f"Пропущено (без фото)          : {stats['skip_photo']}")
    print(f"Пропущено (наявність=0)       : {stats['skip_avail']}")
    print(f"Пропущено (OEM пустий/ціна≤1) : {stats['skip_filter']}")
    print(f"Згенеровано товарів           : {stats['generated']}")
    print(f"Збережено → {out_path}")


if __name__ == "__main__":
    main()
