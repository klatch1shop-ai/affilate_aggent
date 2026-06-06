"""
Prom.ua Feed Validator
Перевіряє файл товарів на відповідність вимогам Prom.ua
"""
import sys, os, re
import openpyxl
sys.path.insert(0, '/home/tekken/agent-system')
from shared.utils.db import get_connection
import logging; logger = logging.getLogger(__name__); logging.basicConfig(level=logging.INFO, format="%(message)s")

# Дозволені одиниці виміру
VALID_UNITS = {
    'шт.', 'т', 'кг', 'г', 'куб.м', 'л', 'кв.м', 'кв.см', 'кв.фут',
    'кв.дм', 'м', 'км', 'дав', 'мішок', 'пара', 'чол.', 'упаковка',
    'сотка', 'пог. м', 'ящик', 'мм', 'мл', 'гр/кв.м', 'кг/кв.м',
    '100 г', 'комплект', 'набір', 'моток', 'рулон', 'послуга', 'см',
    'секція', 'бухта', 'об\'єкт', 'сторінка', 'т/км', 'добу', 'ват',
    'лист', 'карат', 'хвилина', 'кВт', 'мВт', 'бобіна', 'палетомісць',
    'зміна', 'од.', 'година', 'день', 'тиждень', 'місяць'
}

VALID_CURRENCIES = {'UAH', 'USD', 'EUR', 'CHF', 'GBP', 'JPY', 'PLZ', 'BYN', 'KZT', 'MDL'}

VALID_AVAILABILITY = {'+', '-', '!'}

FORBIDDEN_WORDS = [
    'акція', 'безкоштовна доставка', 'знижка', 'найкраща ціна',
    'купити', 'замовити', 'prom.ua', 'уапром'
]

VALID_PRODUCT_TYPES = {'r', 'w', 'u', 's'}


def get_prom_categories():
    """Завантажити всі category_id з БД."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('SELECT category_id, full_path FROM prom_categories')
        result = {row['category_id']: row['full_path'] for row in cur.fetchall()}
        cur.close(); conn.close()
        return result
    except Exception as e:
        logger.error(f'DB error: {e}')
        return {}


def validate_product(row_num, row, categories):
    """Валідує один рядок товару. Повертає список помилок."""
    errors = []
    warnings = []

    def err(field, msg, critical=True):
        if critical:
            errors.append({'field': field, 'message': msg, 'type': 'CRITICAL'})
        else:
            warnings.append({'field': field, 'message': msg, 'type': 'WARNING'})

    # 1. ID товару
    product_id = row.get('Код_товара') or row.get('Ідентифікатор_товару')
    if not product_id:
        err('Код_товара', 'ID товару обов\'язковий')

    # 2. Назва
    name = str(row.get('Название_позиции') or row.get('Назва_позиції') or '').strip()
    if not name:
        err('Название_позиции', 'Назва товару обов\'язкова')
    elif len(name) > 130:
        err('Название_позиции', f'Назва задовга: {len(name)} символів (макс 130)', critical=False)
    elif re.match(r'^\d+$', name):
        err('Название_позиции', 'Назва не може складатися тільки з цифр')
    elif name == name.upper() and len(name) > 3:
        err('Название_позиции', 'Назва написана ВЕЛИКИМИ ЛІТЕРАМИ', critical=False)
    else:
        name_lower = name.lower()
        for word in FORBIDDEN_WORDS:
            if word in name_lower:
                err('Название_позиции', f'Заборонене слово в назві: "{word}"', critical=False)

    # 3. Опис
    desc = str(row.get('Описание') or row.get('Опис') or '').strip()
    if not desc:
        err('Описание', 'Опис товару обов\'язковий', critical=False)
    elif len(desc) < 30:
        err('Описание', f'Опис занадто короткий: {len(desc)} символів (мін 30)', critical=False)
    elif len(desc) > 12160:
        err('Описание', f'Опис задовгий: {len(desc)} символів (макс 12160)', critical=False)

    # 4. Ціна
    price = row.get('Цена') or row.get('Ціна')
    try:
        price_val = float(str(price).replace(',', '.')) if price else 0
        if price_val <= 0:
            err('Цена', 'Ціна повинна бути більше 0')
        elif price_val > 9999999999:
            err('Цена', 'Ціна перевищує максимум')
    except:
        err('Цена', f'Некоректне значення ціни: {price}')

    # 5. Наявність
    avail = str(row.get('Цена от') or row.get('Наявність') or '').strip()
    if avail and avail not in VALID_AVAILABILITY:
        try:
            int(avail)  # може бути кількість днів
        except:
            err('Наявність', f'Некоректне значення наявності: {avail}', critical=False)

    # 6. Одиниця виміру
    unit = str(row.get('Единица_измерения') or row.get('Одиниця_виміру') or '').strip()
    if unit and unit not in VALID_UNITS:
        err('Единица_измерения', f'Невідома одиниця виміру: "{unit}"', critical=False)

    # 7. Валюта
    currency = str(row.get('Валюта') or 'UAH').strip()
    if currency not in VALID_CURRENCIES:
        err('Валюта', f'Невідома валюта: "{currency}"')

    # 8. Категорія
    cat_id = row.get('Ідентифікатор_підрозділу') or row.get('portal_category_id')
    if cat_id:
        try:
            cat_id_int = int(float(str(cat_id)))
            if cat_id_int not in categories:
                err('Ідентифікатор_підрозділу', f'Категорія ID {cat_id_int} не знайдена в Prom')
        except:
            err('Ідентифікатор_підрозділу', f'Некоректний ID категорії: {cat_id}')
    else:
        err('Ідентифікатор_підрозділу', 'Категорія не вказана — буде призначена автоматично', critical=False)

    # 9. Фото
    images = str(row.get('Ссылка_изображения') or row.get('Посилання_зображення') or '').strip()
    if not images:
        err('Ссылка_изображения', 'Немає посилань на фото', critical=False)
    else:
        img_list = [x.strip() for x in images.split(',')]
        if len(img_list) > 10:
            err('Ссылка_изображения', f'Забагато фото: {len(img_list)} (макс 10)', critical=False)
        for img in img_list:
            if not img.startswith('http'):
                err('Ссылка_изображення', f'Невірне посилання на фото: {img}', critical=False)

    return errors, warnings


def validate_file(filepath):
    """Головна функція валідації файлу."""
    logger.info(f'Валідація файлу: {filepath}')
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Знаходимо лист з товарами
    product_sheet = None
    for name in wb.sheetnames:
        if 'product' in name.lower() or 'товар' in name.lower() or 'export' in name.lower():
            product_sheet = wb[name]
            break
    if not product_sheet:
        product_sheet = wb.active

    logger.info(f'Лист: {product_sheet.title}')

    # Читаємо заголовки
    headers = []
    for cell in product_sheet[1]:
        headers.append(str(cell.value or '').strip())
    logger.info(f'Колонки: {headers[:10]}...')

    # Завантажуємо категорії
    categories = get_prom_categories()
    logger.info(f'Категорій в БД: {len(categories)}')

    # Валідуємо кожен рядок
    total = 0
    critical_count = 0
    warning_count = 0
    all_errors = []

    for row_num, row in enumerate(product_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        errors, warnings = validate_product(row_num, row_dict, categories)
        
        total += 1
        critical_count += len(errors)
        warning_count += len(warnings)
        
        for e in errors + warnings:
            all_errors.append({
                'row': row_num,
                'product_id': row_dict.get('Код_товара') or row_dict.get('Ідентифікатор_товару', '?'),
                'name': str(row_dict.get('Название_позиции') or row_dict.get('Назва_позиції', ''))[:50],
                **e
            })

    # Звіт
    print(f'\n{"="*60}')
    print(f'ЗВІТ ВАЛІДАЦІЇ: {os.path.basename(filepath)}')
    print(f'{"="*60}')
    print(f'Всього товарів: {total}')
    print(f'Критичних помилок: {critical_count}')
    print(f'Попереджень: {warning_count}')
    print(f'{"="*60}\n')

    if all_errors:
        print('Перші 20 помилок:')
        for e in all_errors[:20]:
            icon = '❌' if e['type'] == 'CRITICAL' else '⚠️'
            print(f'{icon} Рядок {e["row"]} | {e["name"][:30]} | {e["field"]}: {e["message"]}')

    # Зберігаємо в БД
    save_to_db(os.path.basename(filepath), all_errors)
    
    return total, critical_count, warning_count


def save_to_db(filename, errors):
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM prom_feed_validator_log WHERE file_name = %s', (filename,))
        for e in errors:
            cur.execute('''
                INSERT INTO prom_feed_validator_log 
                (file_name, row_number, product_id, product_name, error_type, error_field, error_message)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (filename, e['row'], str(e['product_id']), e['name'],
                  e['type'], e['field'], e['message']))
        conn.commit()
        cur.close(); conn.close()
        logger.info(f'Збережено {len(errors)} записів в БД')
    except Exception as ex:
        logger.error(f'DB save error: {ex}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Використання: python3 validator.py <шлях_до_файлу>')
        sys.exit(1)
    validate_file(sys.argv[1])
