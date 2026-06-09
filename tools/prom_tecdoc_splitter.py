#!/usr/bin/env python3
"""
tools/prom_tecdoc_splitter.py — розподіл товарів Prom.ua на tecdoc / manual.

Запуск:
  python3 tools/prom_tecdoc_splitter.py <файл.xlsx>
  python3 tools/prom_tecdoc_splitter.py <файл.xlsx> --col-id "Артикул" --col-brand "Бренд"
"""

import argparse
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

ARTICLE_PATTERN = re.compile(r'^[A-Za-z0-9][A-Za-z0-9\-_./ ]{2,18}[A-Za-z0-9]$')

ARTICLE_COL_NAMES = {"артикул", "код", "oem", "part number", "part_number", "номер деталі", "номер детали"}
BRAND_COL_NAMES   = {"виробник", "brand", "бренд", "производитель"}
NO_BRAND_VALUES   = {"", "без бренда", "без бренду", "no brand", "—", "-", "none"}


def detect_article_col(headers: list[str], rows: list[list]) -> str | None:
    """Повертає назву колонки-кандидата для артикула."""
    for h in headers:
        if h.lower().strip() in ARTICLE_COL_NAMES:
            return h

    # евристика: >50% значень колонки відповідають патерну
    col_idx = {h: i for i, h in enumerate(headers)}
    for h, idx in col_idx.items():
        vals = [str(r[idx]).strip() for r in rows if idx < len(r) and r[idx]]
        if not vals:
            continue
        matches = sum(1 for v in vals if ARTICLE_PATTERN.match(v))
        if matches / len(vals) > 0.5:
            return h
    return None


def detect_brand_col(headers: list[str]) -> str | None:
    for h in headers:
        if h.lower().strip() in BRAND_COL_NAMES:
            return h
    return None


def read_xlsx(path: str) -> tuple[list[str], list[list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [list(r) for r in rows[1:] if any(c for c in r)]
    return headers, data


def classify(headers: list[str], rows: list[list],
             col_id: str | None, col_brand: str | None
             ) -> tuple[list[list], list[list]]:

    art_col   = col_id    or detect_article_col(headers, rows)
    brand_col = col_brand or detect_brand_col(headers)

    print(f"  Колонка артикула : {art_col or '(не знайдено)'}")
    print(f"  Колонка бренду   : {brand_col or '(не знайдено)'}")

    art_idx   = headers.index(art_col)   if art_col   and art_col   in headers else None
    brand_idx = headers.index(brand_col) if brand_col and brand_col in headers else None

    tecdoc, manual = [], []
    for row in rows:
        has_art   = False
        has_brand = False

        if art_idx is not None and art_idx < len(row):
            val = str(row[art_idx]).strip()
            has_art = bool(val and val not in ("None", "-", "—"))

        if brand_idx is not None and brand_idx < len(row):
            val = str(row[brand_idx]).strip().lower()
            has_brand = bool(val and val not in NO_BRAND_VALUES)

        if has_art and has_brand:
            tecdoc.append(row)
        else:
            manual.append(row)

    return tecdoc, manual


def save_xlsx(path: str, headers: list[str], rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def top_brands(headers: list[str], rows: list[list], col_brand: str | None, n: int = 5):
    if not col_brand or col_brand not in headers:
        return []
    idx = headers.index(col_brand)
    counts: dict[str, int] = {}
    for row in rows:
        if idx < len(row):
            val = str(row[idx]).strip()
            if val and val.lower() not in NO_BRAND_VALUES and val != "None":
                counts[val] = counts.get(val, 0) + 1
    return sorted(counts.items(), key=lambda x: -x[1])[:n]


def main():
    parser = argparse.ArgumentParser(description="Prom TecDoc splitter")
    parser.add_argument("xlsx", help="Вхідний XLSX файл")
    parser.add_argument("--col-id",    default=None, help="Назва колонки артикула")
    parser.add_argument("--col-brand", default=None, help="Назва колонки бренду")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"Файл не знайдено: {args.xlsx}")

    print(f"\nЧитаємо: {args.xlsx}")
    headers, rows = read_xlsx(args.xlsx)
    if not headers:
        sys.exit("Файл порожній або пошкоджений")

    print(f"Колонки ({len(headers)}): {', '.join(h for h in headers if h)}\n")

    tecdoc, manual = classify(headers, rows, args.col_id, args.col_brand)

    total = len(tecdoc) + len(manual)
    print(f"\n{'─'*50}")
    print(f"Всього товарів    : {total}")
    print(f"З TecDoc кодами   : {len(tecdoc)}")
    if tecdoc:
        print("  Перші 5 прикладів:")
        for r in tecdoc[:5]:
            print(f"    {r}")
    print(f"Ручна сумісність  : {len(manual)}")
    if manual:
        print("  Перші 5 прикладів:")
        for r in manual[:5]:
            print(f"    {r}")

    brand_col = args.col_brand or detect_brand_col(headers)
    brands = top_brands(headers, tecdoc + manual, brand_col)
    if brands:
        print(f"\nТоп-5 виробників (TecDoc):")
        for name, cnt in top_brands(headers, tecdoc, brand_col):
            print(f"  {cnt:>5}  {name}")

    out_dir = Path(args.xlsx).parent / "exports"
    out_dir.mkdir(exist_ok=True)

    tecdoc_path = str(out_dir / "tecdoc_items.xlsx")
    manual_path = str(out_dir / "manual_items.xlsx")
    save_xlsx(tecdoc_path, headers, tecdoc)
    save_xlsx(manual_path, headers, manual)

    print(f"\nЗбережено:")
    print(f"  {tecdoc_path}  ({len(tecdoc)} рядків)")
    print(f"  {manual_path}  ({len(manual)} рядків)")


if __name__ == "__main__":
    main()
