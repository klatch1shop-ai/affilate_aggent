"""
tools/epicentr_fill_attributes.py
Заповнює порожні * колонки xlsx-експорту Єпіцентру.
НЕ додає нових колонок — тільки ті, що вже є у файлі.

Використання: python3 tools/epicentr_fill_attributes.py path/to/export.xlsx
"""

import sys, re, os, shutil
import openpyxl


# ── EAN-13 ────────────────────────────────────────────────────────────────────

def ean13_from_article(article: str) -> str:
    digits = re.sub(r'\D', '', str(article))
    digits = digits[-12:].zfill(12) if len(digits) >= 12 else digits.zfill(12)
    odd   = sum(int(digits[i]) for i in range(0, 12, 2))
    even  = sum(int(digits[i]) for i in range(1, 12, 2))
    return digits + str((10 - (odd + even * 3) % 10) % 10)


# ── Extraction helpers ────────────────────────────────────────────────────────

def extract_model_code(name: str) -> str:
    codes = re.findall(r'\b[A-Z]{2,}[A-Z0-9]*[0-9]+[A-Z0-9]*\b', str(name))
    return codes[-1] if codes else ''


def first_num(text: str):
    m = re.search(r'(\d+(?:[.,]\d+)?)', str(text))
    return float(m.group(1).replace(',', '.')) if m else None


def count_before_unit(text: str):
    m = re.search(r'(\d+)\s*(?:ед|шт|пр|од)\.?\b', str(text), re.I)
    return int(m.group(1)) if m else None


def num_with_unit(text: str, unit: str):
    m = re.search(rf'(\d+(?:[.,]\d+)?)\s*{re.escape(unit)}\b', str(text), re.I)
    return float(m.group(1).replace(',', '.')) if m else None


def drive_square(text: str) -> str:
    for frac in ('3/4', '1/2', '3/8', '1/4'):
        if frac in text:
            return frac + '"'
    if re.search(r'\b1\s*["″]', text):
        return '1"'
    return ''


# ── Category handlers ─────────────────────────────────────────────────────────
# Each returns list of (col_regex_pattern, value).
# Pattern is matched against existing column headers that contain *.
# The regex ^ anchors to the start to avoid false matches.

def h_vorotky(name: str) -> list:
    nl = name.lower()
    if   'набір головок' in nl or ('набір' in nl and 'головк' in nl): vid = 'набір головок'
    elif re.search(r'тріскачк|трещотк', nl):   vid = 'тріскачка'
    elif 'вороток' in nl:                        vid = 'вороток'
    elif 'кардан'  in nl:                        vid = 'кардан'
    elif re.search(r'подовжувач|удлинитель|подовж', nl): vid = 'планка'
    elif re.search(r'перехідник|переходник', nl): vid = 'перехідник'
    elif 'головка' in nl:                         vid = 'головка'
    else:                                          vid = 'головка'
    return [(r'Вид\s*\*', vid)]


def h_klyuchi(name: str) -> list:
    nl = name.lower()
    if   re.search(r'ріжков', nl):                     vyd = 'ріжковий'
    elif re.search(r'накидн', nl):                     vyd = 'накидний'
    elif re.search(r'комбінован', nl):                 vyd = 'комбінований'
    elif re.search(r'hex|torx|імбус|шестигранн', nl):  vyd = 'шестигранний (імбусовий)'
    elif re.search(r'динамометричн', nl):              vyd = 'динамометричний'
    elif re.search(r'розвідн|разводн', nl):            vyd = ''   # немає у довіднику 903
    elif re.search(r'трубн', nl):                      vyd = 'трубний'
    else:                                               vyd = 'ріжковий'
    typ = 'набір ключів' if 'набір' in nl else 'ключ'
    qty = count_before_unit(name) or ''
    # Розмір — multiselect, значення тільки з довідника; залишаємо порожнім
    return [
        (r'Вид ключів\s*\*',         vyd),
        (r'Тип\s*\*',                typ),
        (r'Кількість у наборі\s*\*', qty),
    ]


def h_vikrutky(name: str) -> list:
    nu, nl = name.upper(), name.lower()
    if   re.search(r'\bPZ\b|pozidriv', nu):        shlits = 'позидрів (PZ)'
    elif re.search(r'\bPH\b|phillips', nu):         shlits = 'хрестоподібний (PH)'
    elif re.search(r'TORX|T\d+', nu):              shlits = 'зірочка (TORX)'
    elif re.search(r'\bHEX\b', nu):                shlits = 'шестигранний (HEX)'
    else:                                            shlits = 'прямий (SL)'
    vyd = 'набір викруток' if 'набір' in nl else 'викрутка'
    return [
        (r'Вид викрутки\s*\*', vyd),
        (r'Тип шліца\s*\*',    shlits),
    ]


def h_sharnirno(name: str) -> list:
    size_mm = num_with_unit(name, 'мм') or ''
    size_in = drive_square(name)
    return [
        (r'Типовий розмір\s*\*.*мм',  size_mm),
        (r'Типовий розмір\s*\*.*"',   size_in),
    ]


def h_bity(name: str) -> list:
    nu, nl = name.upper(), name.lower()
    qty = count_before_unit(name) or 1
    sm  = re.search(r'\b(PH\d+|PZ\d+|T\d+|HEX[\d.]+|SL[\d.]+)\b', nu)
    size = sm.group(1) if sm else ''
    if   re.search(r'\bPZ\b', nu):    typ = 'позидрів (PZ)'
    elif re.search(r'\bPH\b', nu):    typ = 'хрестоподібна (PH)'
    elif re.search(r'TORX|T\d+', nu): typ = 'зірочка (TORX)'
    elif re.search(r'\bHEX\b', nu):   typ = 'шестигранна (HEX)'
    elif re.search(r'\bSL\b', nu):    typ = 'пряма (SL)'
    else:                              typ = 'хрестоподібна (PH)'
    length = num_with_unit(name, 'мм') or 25
    return [
        (r'Матеріал виробу\s*\*',  'хромованадієва сталь'),
        (r'Розмір біти\s*\*',      size),
        (r'Тип біти\s*\*',         typ),
        (r'Довжина біти\s*\*',     length),
        (r'Кількість\s*\*',        qty),          # 'Кількість (float) (шт.)'
    ]


def h_molotky(name: str) -> list:
    nl = name.lower()
    if   re.search(r'рихтув', nl):               func = 'рихтувальний'
    elif re.search(r'мідн', nl):                 func = 'мідний'
    elif re.search(r'гумов', nl):                func = 'гумовий'
    elif re.search(r'полімерн|пластик|нейлон', nl): func = 'полімерний'
    else:                                          func = 'слюсарний'
    vyd = 'набір молотків' if 'набір' in nl else 'молоток'
    wt  = num_with_unit(name, 'г')
    if not wt:
        kg = num_with_unit(name, 'кг')
        wt = kg * 1000 if kg else 500
    return [
        (r'Вид\s*\*',         vyd),
        (r'Вид молотків\s*',  func),   # optional — no *, but try anyway
        (r'(?:^|\s)Вага\s*\*', wt),   # 'Вага* (float) (г)', not 'Вага упаковки*'
    ]


def h_dynamometr(name: str) -> list:
    rm = re.search(r'(\d+(?:[.,]\d+)?)\s*[-–]\s*(\d+(?:[.,]\d+)?)', name)
    if rm:
        mn = float(rm.group(1).replace(',', '.'))
        mx = float(rm.group(2).replace(',', '.'))
    else:
        nm = re.search(r'(\d+(?:[.,]\d+)?)\s*[HН][·•]?[мm]', name)
        mx = float(nm.group(1).replace(',', '.')) if nm else 100
        mn = 5
    return [
        (r'Мінімальне зусилля\s*\*',    mn),
        (r'Максимальне зусилля\s*\*',   mx),
    ]


def h_pnevmo_nabir(name: str) -> list:
    nl  = name.lower()
    if   'гайковерт' in nl: pred = 'гайковерт пневматичний'
    elif 'шліфмаш'   in nl: pred = 'пневматична шліфмашина'
    elif 'дриль'     in nl: pred = 'пневмодриль'
    else:                    pred = 'пневматичний інструмент'
    return [(r'Предмети в наборі\s*\*', pred)]


def h_nabory(name: str) -> list:
    nl  = name.lower()
    qty = count_before_unit(name) or first_num(name) or 1
    n   = int(qty)
    sel = 'до 20' if n < 20 else '20-49' if n < 50 else '50-110' if n <= 110 else 'більше 110'
    pack  = 'валіза пластикова' if re.search(r'валіз|кейс', nl) else 'коробка картонна'
    sfera = 'автомобільний' if re.search(r'авто|сто', nl) else 'універсальний'
    return [
        (r'Кількість в наборі\s*\*',   sel),
        (r'^Упаковка\s*\*',            pack),
        (r'Кількість у наборі\s*\*',   qty),
        (r'Сфера застосування\s*\*',   sfera),
    ]


def h_testery(name: str) -> list:
    nl = name.lower()
    vyd = 'мультиметр' if 'мультиметр' in nl else \
          'індикатор фази' if 'індикатор' in nl else 'тестер електричний'
    return [(r'Вид\s*\*', vyd)]


def h_multymetry(_: str) -> list:
    return [
        (r'Тип живлення\s*\*',          'батарейки'),
        (r'Вимірювання та тести\s*\*',  'напруга'),
    ]


def h_domkraty(_: str) -> list:
    return []   # Домкрати файл не має специфічних * атрибутів


def h_farbopulty(name: str) -> list:
    diam = num_with_unit(name, 'мм') or 1.4
    return [
        (r'Діаметр сопла\s*\*',     diam),
        (r'Витрата повітря\s*\*',   200),
    ]


HANDLERS = {
    'Воротки, тріскачки та головки': h_vorotky,
    'Ключі та набори ключів':        h_klyuchi,
    'Динамометричні ключі':          h_dynamometr,
    'Викрутки':                      h_vikrutky,
    'Шарнірно-губцевий інструмент':  h_sharnirno,
    'Біти для шуруповерта':          h_bity,
    'Молотки':                       h_molotky,
    'Набори пневмоінструменту':      h_pnevmo_nabir,
    'Набори інструментів':           h_nabory,
    'Тестери електричні':            h_testery,
    'Мультиметри':                   h_multymetry,
    'Домкрати':                      h_domkraty,
    'Фарбопульти пневматичні':       h_farbopulty,
}


# ── Column index helpers ──────────────────────────────────────────────────────

def build_col_index(headers: list) -> dict:
    """Returns {header_string: 0-based-index} for all non-None headers."""
    return {str(h): i for i, h in enumerate(headers) if h is not None}


def find_col(headers: list, pattern: str, require_star: bool = True) -> int | None:
    """First column whose header matches pattern (optionally must contain *)."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        hs = str(h)
        if require_star and '*' not in hs:
            continue
        if re.search(pattern, hs, re.I | re.S):
            return i
    return None


def is_empty(value) -> bool:
    return value in (None, '', '0', 0)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('Usage: python3 tools/epicentr_fill_attributes.py path/to/export.xlsx')
        sys.exit(1)

    in_path  = sys.argv[1]
    base_dir = os.path.dirname(in_path) or '.'
    base_nam = os.path.splitext(os.path.basename(in_path))[0]
    out_path = os.path.join(base_dir, base_nam + '_filled.xlsx')

    # ── Step 1: copy to preserve structure ───────────────────────────────────
    shutil.copy(in_path, out_path)
    print(f'Копію збережено: {out_path}')

    # ── Step 2: load copy ─────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data_rows = ws.max_row - 1
    print(f'Колонок: {len(headers)}  Рядків: {data_rows}')

    # Detect primary category from first data row
    cat_ci = find_col(headers, r'Основна категорія', require_star=False)
    first_cat = ws.cell(row=2, column=(cat_ci or 0) + 1).value if cat_ci is not None else '?'
    print(f'Перша категорія: {first_cat}')

    # Pre-locate fixed columns (0-based)
    name_ci    = find_col(headers, r'Назва товару.*ua',     require_star=False)
    article_ci = find_col(headers, r'Артикул',              require_star=True)
    barcode_ci = find_col(headers, r'Штрих код',            require_star=False)
    model_ci   = find_col(headers, r'Модель виробника',     require_star=False)
    pack_cis   = {
        'Висота упаковки':  find_col(headers, r'Висота упаковки'),
        'Глибина упаковки': find_col(headers, r'Глибина упаковки'),
        'Вага упаковки':    find_col(headers, r'Вага упаковки'),
        'Ширина упаковки':  find_col(headers, r'Ширина упаковки'),
    }

    stats: dict[str, int] = {}

    def get(ri: int, ci: int):
        return ws.cell(row=ri, column=ci + 1).value

    def put(ri: int, ci: int, value, key: str):
        ws.cell(row=ri, column=ci + 1, value=value)
        stats[key] = stats.get(key, 0) + 1

    # ── Step 3: process rows ──────────────────────────────────────────────────
    for ri in range(2, ws.max_row + 1):
        name    = str(get(ri, name_ci)    or '') if name_ci    is not None else ''
        article = str(get(ri, article_ci) or '') if article_ci is not None else ''
        cat     = str(get(ri, cat_ci)     or '') if cat_ci     is not None else ''

        # Packaging dimensions: 0 → 1
        for pname, pc in pack_cis.items():
            if pc is not None and is_empty(get(ri, pc)):
                put(ri, pc, 1, pname)

        # Barcode: generate EAN-13
        if barcode_ci is not None and is_empty(get(ri, barcode_ci)) and article:
            put(ri, barcode_ci, ean13_from_article(article), 'Штрих код')

        # Model code: extract from name
        if model_ci is not None and is_empty(get(ri, model_ci)) and name:
            code = extract_model_code(name)
            if code:
                put(ri, model_ci, code, 'Модель виробника')

        # Category-specific attributes
        handler = HANDLERS.get(cat)
        if not handler:
            continue
        for pattern, value in handler(name):
            if not value and value != 0:
                continue
            ci = find_col(headers, pattern, require_star=False)
            if ci is not None and is_empty(get(ri, ci)):
                # Extract human-readable key: take text before first regex operator
                key = re.split(r'[\\*+?^${}|()\[\]]', pattern)[0].strip()
                put(ri, ci, value, key)

    # ── Step 4: save in place ─────────────────────────────────────────────────
    wb.save(out_path)
    print(f'Збережено: {out_path}')

    # ── Statistics ────────────────────────────────────────────────────────────
    print(f'\n{"Поле":<55} {"Заповнено":>10}')
    print('─' * 67)
    generic = ['Висота упаковки', 'Глибина упаковки', 'Вага упаковки', 'Ширина упаковки',
               'Штрих код', 'Модель виробника']
    for k in generic:
        if k in stats:
            print(f'  {k:<53} {stats[k]:>10}')
    cat_stats = {k: v for k, v in stats.items() if k not in generic}
    if cat_stats:
        print('  ' + '─' * 65)
        for k, v in sorted(cat_stats.items(), key=lambda x: -x[1]):
            print(f'  {k:<53} {v:>10}')
    print('─' * 67)
    print(f'  {"ВСЬОГО":<53} {sum(stats.values()):>10}')


if __name__ == '__main__':
    main()
