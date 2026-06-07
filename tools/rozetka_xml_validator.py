#!/usr/bin/env python3
"""
tools/rozetka_xml_validator.py
Universal validator for Rozetka seller XML (YML) feeds.

Sources:
  sellerhelp.rozetka.com.ua/p185-pricelist-requirements.html
  shared/knowledge_base/rozetka/xml_faq.txt
  docs/XML_VALIDATION_PLAN.md

Usage:
  python3 tools/rozetka_xml_validator.py [path_to_xml]
  python3 tools/rozetka_xml_validator.py data/katran_rozetka.xml
"""

import os
import re
import sys
import json
from collections import defaultdict
from datetime import datetime
from xml.etree import ElementTree as ET

# ─── constants ───────────────────────────────────────────────────────────────

DEFAULT_CATEGORY_ID = "25636737"   # Ручний інструмент (наш fallback — попередження)
MAX_URL_LEN         = 1999
MAX_NAME_LEN        = 255
WARN_NAME_LEN       = 150
MAX_ARTICLE_LEN     = 255
MIN_PARAMS          = 3
CYRILLIC_RE         = re.compile(r"[а-яА-ЯіІєЄїЇёЁ]")
URL_RE              = re.compile(r"^https?://", re.IGNORECASE)
HTTPS_RE            = re.compile(r"^https://", re.IGNORECASE)
FORBIDDEN_NAME_START = re.compile(r"^[!@#$%^&*()\[\]{}<>|\\/'\"~`]")
DIGIT_START_RE      = re.compile(r"^\d")

VENDOR_BLOCKLIST = {"no name", "noname", "no-name", "unknown", "без назви", ""}

# ─── issue model ─────────────────────────────────────────────────────────────

class Issue:
    __slots__ = ("level", "offer_id", "field", "message")

    def __init__(self, level: str, offer_id: str, field: str, message: str):
        self.level    = level     # "ERROR" | "WARN"
        self.offer_id = offer_id
        self.field    = field
        self.message  = message

    def to_dict(self) -> dict:
        return {
            "level":    self.level,
            "offer_id": self.offer_id,
            "field":    self.field,
            "message":  self.message,
        }


# ─── helpers ─────────────────────────────────────────────────────────────────

def _txt(el, tag: str, default: str = "") -> str:
    child = el.find(tag)
    if child is None or child.text is None:
        return default
    return child.text.strip()


def _all_txt(el, tag: str) -> list[str]:
    return [c.text.strip() for c in el.findall(tag) if c.text]


# ─── validator ───────────────────────────────────────────────────────────────

class RozetkaXMLValidator:
    def __init__(self, xml_path: str):
        self.xml_path   = xml_path
        self.issues: list[Issue] = []
        self.offers_data: list[dict] = []
        self.category_ids: set[str] = set()
        self.category_names: dict[str, str] = {}
        self.parse_ok  = False
        self.root      = None

    # ── error / warn shortcuts ────────────────────────────────────────────────

    def _err(self, offer_id: str, field: str, msg: str):
        self.issues.append(Issue("ERROR", offer_id, field, msg))

    def _warn(self, offer_id: str, field: str, msg: str):
        self.issues.append(Issue("WARN",  offer_id, field, msg))

    # ── 1. structural checks ─────────────────────────────────────────────────

    def _check_structure(self) -> bool:
        # 1a. XML syntax + UTF-8
        try:
            with open(self.xml_path, "rb") as f:
                raw = f.read()
        except OSError as e:
            self._err("FILE", "file", f"Не вдалось прочитати файл: {e}")
            return False

        if b"encoding" in raw[:200].lower() and b"utf-8" not in raw[:200].lower():
            self._err("FILE", "encoding", "XML декларує не UTF-8 кодування")

        try:
            self.root = ET.fromstring(raw)
        except ET.ParseError as e:
            self._err("FILE", "xml_syntax", f"Невалідний XML синтаксис: {e}")
            return False

        # 1b. Root tag
        if self.root.tag != "yml_catalog":
            self._err("FILE", "root", f"Кореневий тег має бути yml_catalog, знайдено: {self.root.tag}")
            return False

        # 1c. Required containers
        shop = self.root.find("shop")
        if shop is None:
            self._err("FILE", "shop", "Відсутній тег <shop>")
            return False

        for required in ("currencies", "categories", "offers"):
            if shop.find(required) is None:
                self._err("FILE", required, f"Відсутній тег <{required}>")

        return True

    def _load_categories(self):
        shop = self.root.find("shop")
        if shop is None:
            return
        cats_el = shop.find("categories")
        if cats_el is None:
            return
        for cat in cats_el.findall("category"):
            cid = cat.get("id", "").strip()
            name = (cat.text or "").strip()
            if cid:
                self.category_ids.add(cid)
                self.category_names[cid] = name

    # ── 2. per-offer checks ───────────────────────────────────────────────────

    def _check_offer(self, offer: ET.Element, seen_ids: set, name_counter: dict):
        oid = offer.get("id", "").strip()
        if not oid:
            self._err("?", "offer_id", "Оффер без атрибуту id")
            oid = "NOID"
        elif oid in seen_ids:
            self._err(oid, "offer_id", f"Дублікат offer id: {oid}")
        seen_ids.add(oid)

        result = {"id": oid, "errors": 0, "warns": 0}

        # price
        price_txt = _txt(offer, "price")
        try:
            price = float(price_txt)
            if price <= 0:
                self._err(oid, "price", f"price <= 0: {price}")
                result["errors"] += 1
        except ValueError:
            self._err(oid, "price", f"price не є числом: '{price_txt}'")
            result["errors"] += 1

        # currencyId
        currency = _txt(offer, "currencyId")
        if currency != "UAH":
            self._err(oid, "currencyId", f"currencyId має бути UAH, знайдено: '{currency}'")
            result["errors"] += 1

        # categoryId
        cat_id = _txt(offer, "categoryId")
        if not cat_id:
            self._err(oid, "categoryId", "Відсутній тег <categoryId>")
            result["errors"] += 1
        elif cat_id not in self.category_ids:
            self._err(oid, "categoryId", f"categoryId '{cat_id}' не оголошено в <categories>")
            result["errors"] += 1
        elif cat_id == DEFAULT_CATEGORY_ID:
            self._warn(oid, "categoryId", f"categoryId = {DEFAULT_CATEGORY_ID} (DEFAULT — 'Ручний інструмент'). Товар може бути відхилений.")
            result["warns"] += 1

        # pictures
        pics = _all_txt(offer, "picture")
        if not pics:
            self._err(oid, "picture", "Відсутнє хоча б одне фото <picture>")
            result["errors"] += 1
        for pic in pics:
            if not URL_RE.match(pic):
                self._err(oid, "picture", f"URL фото не починається з http: {pic[:80]}")
                result["errors"] += 1
            elif not HTTPS_RE.match(pic):
                self._warn(oid, "picture", f"URL фото HTTP замість HTTPS: {pic[:80]}")
                result["warns"] += 1
            if CYRILLIC_RE.search(pic):
                self._err(oid, "picture", f"URL фото містить кирилицю: {pic[:80]}")
                result["errors"] += 1
            if len(pic) > MAX_URL_LEN:
                self._err(oid, "picture", f"URL фото > {MAX_URL_LEN} символів ({len(pic)})")
                result["errors"] += 1

        # name / name_ua
        name    = _txt(offer, "name")
        name_ua = _txt(offer, "name_ua")
        effective_name = name_ua or name
        if not effective_name:
            self._err(oid, "name", "Відсутні тegi <name> і <name_ua>")
            result["errors"] += 1
        else:
            if len(effective_name) < 5:
                self._err(oid, "name", f"Назва коротша 5 символів: '{effective_name}'")
                result["errors"] += 1
            elif len(effective_name) > MAX_NAME_LEN:
                self._err(oid, "name", f"Назва > {MAX_NAME_LEN} символів ({len(effective_name)})")
                result["errors"] += 1
            if FORBIDDEN_NAME_START.match(effective_name):
                self._err(oid, "name", f"Назва починається з забороненого символу: '{effective_name[:30]}'")
                result["errors"] += 1
            elif DIGIT_START_RE.match(effective_name):
                self._warn(oid, "name", f"Назва починається з цифри: '{effective_name[:50]}'")
                result["warns"] += 1
            if len(effective_name) > WARN_NAME_LEN:
                self._warn(oid, "name", f"Назва довша {WARN_NAME_LEN} символів ({len(effective_name)})")
                result["warns"] += 1
            # duplicate name tracking
            name_counter[effective_name] += 1

        # vendor
        vendor = _txt(offer, "vendor")
        if not vendor or vendor.lower() in VENDOR_BLOCKLIST:
            self._err(oid, "vendor", f"vendor відсутній або недопустимий: '{vendor}'")
            result["errors"] += 1
        elif vendor.lower() in ("без бренду", "no brand", "nobrand"):
            self._warn(oid, "vendor", f"vendor = '{vendor}' (товар без бренду — ризик модерації)")
            result["warns"] += 1
        elif URL_RE.match(vendor):
            self._err(oid, "vendor", f"vendor виглядає як URL: '{vendor[:60]}'")
            result["errors"] += 1

        # article
        article = _txt(offer, "article")
        if not article:
            self._warn(oid, "article", "Відсутній тег <article> (рекомендовано)")
            result["warns"] += 1
        elif len(article) > MAX_ARTICLE_LEN:
            self._err(oid, "article", f"article > {MAX_ARTICLE_LEN} символів ({len(article)})")
            result["errors"] += 1
        if article and effective_name and article == effective_name and len(article) < 15:
            self._warn(oid, "name", f"Назва = артикул ('{article}') — рекомендується змістовна назва")
            result["warns"] += 1

        # stock_quantity
        sq_txt = _txt(offer, "stock_quantity")
        try:
            sq = int(float(sq_txt))
            if sq <= 0:
                self._err(oid, "stock_quantity", f"stock_quantity <= 0: {sq}")
                result["errors"] += 1
        except ValueError:
            self._err(oid, "stock_quantity", f"stock_quantity не є числом: '{sq_txt}'")
            result["errors"] += 1

        # params
        params = offer.findall("param")
        param_names = [p.get("name", "") for p in params]
        if len(params) < MIN_PARAMS:
            self._warn(oid, "param", f"Менше {MIN_PARAMS} характеристик <param>: знайдено {len(params)}")
            result["warns"] += 1
        for p in params:
            pname = p.get("name", "")
            pval  = (p.text or "").strip()
            if pname == "Гарантія" and pval in ("0", "0 міс", "0 місяців"):
                self._warn(oid, "param", "Гарантія = 0 міс — не виводити, Розетка не приймає нульову гарантію")
                result["warns"] += 1

        self.offers_data.append({
            "id":       oid,
            "cat_id":   cat_id,
            "name":     effective_name[:100] if effective_name else "",
            "vendor":   vendor,
            "price":    price_txt,
            "pics":     len(pics),
            "params":   len(params),
            "errors":   result["errors"],
            "warns":    result["warns"],
        })

    # ── 3. duplicate names (post-pass) ────────────────────────────────────────

    def _check_dup_names(self, name_counter: dict):
        for name, count in name_counter.items():
            if count > 1:
                self._warn("MULTIPLE", "name", f"Дублікат назви ({count} offerів): '{name[:70]}'")

    # ── 4. main entry ─────────────────────────────────────────────────────────

    def run(self) -> dict:
        print(f"\n{'='*65}")
        print(f"  Rozetka XML Validator")
        print(f"  Файл: {self.xml_path}")
        print(f"  Час : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*65}\n")

        # structural
        print("▶ Перевірка структури XML...")
        ok = self._check_structure()
        if not ok:
            print("  ✗ Критична структурна помилка — подальша перевірка неможлива\n")
            return self._build_report()

        self.parse_ok = True
        self._load_categories()
        print(f"  ✓ XML синтаксис: OK")
        print(f"  ✓ Категорій оголошено: {len(self.category_ids)}")

        # offers
        shop     = self.root.find("shop")
        offers_el = shop.find("offers")
        all_offers = offers_el.findall("offer") if offers_el is not None else []
        total = len(all_offers)
        print(f"  ✓ Офферів у файлі: {total}\n")

        print("▶ Перевірка офферів...")
        seen_ids     = set()
        name_counter = defaultdict(int)
        for offer in all_offers:
            self._check_offer(offer, seen_ids, name_counter)

        self._check_dup_names(name_counter)

        return self._build_report()

    # ── 5. reporting ──────────────────────────────────────────────────────────

    def _build_report(self) -> dict:
        errors = [i for i in self.issues if i.level == "ERROR"]
        warns  = [i for i in self.issues if i.level == "WARN"]
        total  = len(self.offers_data)

        # per-offer counts
        offers_with_error = sum(1 for o in self.offers_data if o["errors"] > 0)
        offers_clean      = total - offers_with_error
        pass_rate         = (offers_clean / total * 100) if total else 0

        # category distribution
        cat_dist = defaultdict(int)
        for o in self.offers_data:
            cat_dist[o["cat_id"]] += 1

        # top-10 worst offers
        worst = sorted(self.offers_data, key=lambda x: -(x["errors"] * 100 + x["warns"]))[:10]

        # issues by offer_id
        issues_by_offer = defaultdict(list)
        for iss in self.issues:
            issues_by_offer[iss.offer_id].append(iss)

        report = {
            "meta": {
                "file":        self.xml_path,
                "generated":   datetime.now().isoformat(),
                "parse_ok":    self.parse_ok,
            },
            "summary": {
                "total_offers":        total,
                "total_errors":        len(errors),
                "total_warnings":      len(warns),
                "offers_with_errors":  offers_with_error,
                "offers_clean":        offers_clean,
                "pass_rate_pct":       round(pass_rate, 1),
                "categories_count":    len(self.category_ids),
            },
            "category_distribution": dict(
                sorted(cat_dist.items(), key=lambda x: -x[1])
            ),
            "top10_worst_offers": worst,
            "all_issues": [i.to_dict() for i in self.issues],
        }

        self._print_report(report, errors, warns)
        return report

    def _print_report(self, report: dict, errors: list, warns: list):
        s = report["summary"]

        # category stats
        cat_lines = []
        for cid, cnt in list(report["category_distribution"].items())[:15]:
            name = self.category_names.get(cid, "?")
            marker = " ⚠ DEFAULT" if cid == DEFAULT_CATEGORY_ID else ""
            cat_lines.append(f"    {cnt:>5}  {cid:<12} {name[:35]}{marker}")

        # errors by field
        err_by_field = defaultdict(int)
        for e in errors:
            err_by_field[e.field] += 1
        warn_by_field = defaultdict(int)
        for w in warns:
            warn_by_field[w.field] += 1

        print(f"{'─'*65}")
        print(f"  ПІДСУМОК")
        print(f"{'─'*65}")
        print(f"  Офферів всього         : {s['total_offers']}")
        print(f"  Помилок (ERROR)        : {s['total_errors']}")
        print(f"  Попереджень (WARN)     : {s['total_warnings']}")
        print(f"  Офферів з помилками    : {s['offers_with_errors']}")
        print(f"  Офферів без помилок    : {s['offers_clean']}")
        print(f"  Відсоток валідних      : {s['pass_rate_pct']}%")
        print()

        if s["total_errors"] > 0:
            print(f"  ПОМИЛКИ по полях:")
            for field, cnt in sorted(err_by_field.items(), key=lambda x: -x[1])[:12]:
                print(f"    ✗ {field:<22} {cnt:>5}")
            print()

        if s["total_warnings"] > 0:
            print(f"  ПОПЕРЕДЖЕННЯ по полях:")
            for field, cnt in sorted(warn_by_field.items(), key=lambda x: -x[1])[:12]:
                print(f"    ⚠ {field:<22} {cnt:>5}")
            print()

        print(f"  РОЗПОДІЛ ПО КАТЕГОРІЯХ (топ-15):")
        for line in cat_lines:
            print(line)
        print()

        if report["top10_worst_offers"]:
            print(f"  ТОП-10 ПРОБЛЕМНИХ ОФФЕРІВ:")
            for o in report["top10_worst_offers"]:
                if o["errors"] == 0 and o["warns"] == 0:
                    break
                print(f"    [{o['id'][:20]:<20}] ERR={o['errors']} WARN={o['warns']}  {o['name'][:40]}")
            print()

        # final verdict
        if s["total_errors"] == 0:
            print(f"  ✅  СТРУКТУРНИХ ПОМИЛОК НЕМАЄ — файл може бути завантажений")
        else:
            print(f"  ✗  ФАЙЛ МАЄ {s['total_errors']} ПОМИЛОК — потребує виправлення")

        print(f"{'='*65}\n")

        # sample errors
        shown = 0
        for iss in errors[:20]:
            print(f"  ERROR [{iss.offer_id[:20]}] {iss.field}: {iss.message[:80]}")
            shown += 1
        if len(errors) > 20:
            print(f"  ... ще {len(errors) - 20} помилок (повний список у JSON/XLSX звіті)")
        if shown:
            print()


# ─── output writers ───────────────────────────────────────────────────────────

def write_json(report: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON звіт: {path}")


def write_xlsx(report: dict, path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠ openpyxl не встановлено — XLSX звіт пропущено")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E4053")
    err_fill  = PatternFill("solid", fgColor="FADBD8")
    warn_fill = PatternFill("solid", fgColor="FEF9E7")
    ok_fill   = PatternFill("solid", fgColor="D5F5E3")

    s = report["summary"]
    rows = [
        ("Параметр", "Значення"),
        ("Файл",               report["meta"]["file"]),
        ("Дата перевірки",     report["meta"]["generated"][:19]),
        ("Офферів всього",     s["total_offers"]),
        ("Помилок (ERROR)",    s["total_errors"]),
        ("Попереджень (WARN)", s["total_warnings"]),
        ("Офферів з помилками",s["offers_with_errors"]),
        ("Офферів валідних",   s["offers_clean"]),
        ("Відсоток валідних",  f"{s['pass_rate_pct']}%"),
        ("Категорій",          s["categories_count"]),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
        if i == 1:
            for c in (ws.cell(i, 1), ws.cell(i, 2)):
                c.font = hdr_font
                c.fill = hdr_fill
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    # ── Sheet 2: Issues ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues")
    hdrs = ["Рівень", "Offer ID", "Поле", "Повідомлення"]
    for col, h in enumerate(hdrs, 1):
        c = ws2.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill

    for row, iss in enumerate(report["all_issues"], 2):
        ws2.cell(row, 1, iss["level"])
        ws2.cell(row, 2, iss["offer_id"])
        ws2.cell(row, 3, iss["field"])
        ws2.cell(row, 4, iss["message"])
        fill = err_fill if iss["level"] == "ERROR" else warn_fill
        for col in range(1, 5):
            ws2.cell(row, col).fill = fill

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 80
    ws2.auto_filter.ref = f"A1:D{len(report['all_issues'])+1}"

    # ── Sheet 3: Categories ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Categories")
    for col, h in enumerate(["rz_id", "Назва категорії", "Офферів", "Примітка"], 1):
        c = ws3.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill

    cat_names_map = {
        k: v for k, v in zip(
            report["category_distribution"].keys(),
            [report["meta"].get("cat_names", {}).get(k, "") for k in report["category_distribution"]]
        )
    }

    for row, (cid, cnt) in enumerate(report["category_distribution"].items(), 2):
        ws3.cell(row, 1, cid)
        ws3.cell(row, 2, report.get("_cat_names", {}).get(cid, ""))
        ws3.cell(row, 3, cnt)
        note = "DEFAULT — ризик відхилення" if cid == DEFAULT_CATEGORY_ID else ""
        ws3.cell(row, 4, note)
        if cid == DEFAULT_CATEGORY_ID:
            for col in range(1, 5):
                ws3.cell(row, col).fill = warn_fill

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 35
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 35

    # ── Sheet 4: Worst Offers ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Worst Offers")
    hdrs4 = ["Offer ID", "Помилок", "Попереджень", "Назва", "Категорія", "Вендор", "Фото", "Параметрів"]
    for col, h in enumerate(hdrs4, 1):
        c = ws4.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill

    for row, o in enumerate(report["top10_worst_offers"], 2):
        ws4.cell(row, 1, o["id"])
        ws4.cell(row, 2, o["errors"])
        ws4.cell(row, 3, o["warns"])
        ws4.cell(row, 4, o["name"])
        ws4.cell(row, 5, o["cat_id"])
        ws4.cell(row, 6, o["vendor"])
        ws4.cell(row, 7, o["pics"])
        ws4.cell(row, 8, o["params"])
        if o["errors"] > 0:
            ws4.cell(row, 2).fill = err_fill
        elif o["warns"] > 0:
            ws4.cell(row, 3).fill = warn_fill

    for col, w in enumerate([20, 8, 12, 45, 14, 20, 6, 10], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(path)
    print(f"  XLSX звіт: {path}")


# ─── entry point ─────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Rozetka XML Validator")
    parser.add_argument("xml_file", nargs="?", default=None, help="XML file to validate")
    parser.add_argument(
        "--no-xlsx", action="store_true",
        help="Skip Excel report (use on servers without AVX/openpyxl support)"
    )
    args = parser.parse_args()

    if args.xml_file:
        xml_path = args.xml_file
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        xml_path = os.path.join(script_dir, "..", "data", "katran_rozetka.xml")

    if not os.path.isabs(xml_path):
        xml_path = os.path.join(os.getcwd(), xml_path)

    xml_path = os.path.normpath(xml_path)

    if not os.path.exists(xml_path):
        print(f"Файл не знайдено: {xml_path}")
        sys.exit(1)

    validator = RozetkaXMLValidator(xml_path)
    report    = validator.run()

    report["_cat_names"] = validator.category_names

    json_path = "/tmp/validation_report.json"
    print("▶ Зберігаю звіти...")
    write_json(report, json_path)
    if not args.no_xlsx:
        write_xlsx(report, "/tmp/validation_report.xlsx")
    print()

    errors = report["summary"]["total_errors"]
    sys.exit(0 if errors == 0 else 1)


if __name__ == "__main__":
    main()
